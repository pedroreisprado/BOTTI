import openpyxl
import time
from datetime import datetime
from dateutil import parser

from config import SHEET_NOVOS_PRODUTOS, SHEET_COMPRAS_CLIENTES, SHEET_VENDEDORES
from functions.logger import logger

def selectProducts():
    file = openpyxl.load_workbook(SHEET_NOVOS_PRODUTOS)
    time.sleep(5)

    sheet = file.active
    data = []

    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(row[i] is not None and row[i] != '' for i in [0, 1, 2, 3, 4, 5, 6, 7]):
                row_data = {
                    "id": str(row[0]),
                    "newProduct": row[1].strip(),
                    "codNewProd": str(row[2]),
                    "codProdRef": str(row[3]),    
                    "productRef": str(row[4]),
                    "monthBuy": int(row[5]),
                    "yearBuy": int(row[6]),
                    "prazo": int(row[7])
                }
                data.append(row_data)

    except Exception as e:
        logger(f"!! O script encontrou um erro (selectProducts) : {e}")

    file.close()

    logger(f"Planilha processada produtos encontrados dentro do prazo: {str(data)}")

    return data

def checkPurchase(data):
    file = openpyxl.load_workbook(SHEET_COMPRAS_CLIENTES, data_only=True)
    time.sleep(10)
    sheet = file.active

    result = []

    try:
        for item in data:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (
                    row[0] is not None and row[0] != '' and
                    row[1] is not None and row[1] != '' and row[1] != 'CONSUMIDOR FINAL LOJA' and
                    row[22] is not None and row[22] != '' and
                    row[69] is not None and row[69] != '' and
                    row[71] is not None and row[71] != '' and
                    str(row[22]) == item['codProdRef'] and
                    datetime.strptime(str(row[14]), '%Y-%m-%d %H:%M:%S').month == item['monthBuy'] and
                    datetime.strptime(str(row[14]), '%Y-%m-%d %H:%M:%S').year == item['yearBuy']
                ):
                    row_data = {
                        "id": item['id'],
                        "codNewProd": item['codNewProd'],
                        "newProduct": item['newProduct'],
                        "codClient": row[0],
                        "nameClient": str(row[1]),
                        "codProduct": str(row[22]),
                        "grid": str(row[69]),    
                        "seller": str(row[71]),
                        "prazo": item['prazo']
                    }
                    result.append(row_data)
                    
    except Exception as e:
        logger(f"!! O script encontrou um erro (checkPurchase) : {e}")

    file.close()

    logger(f"Clientes encontrados: {str(result)}")

    return result

def getHistory(data):
    file = openpyxl.load_workbook(SHEET_COMPRAS_CLIENTES, data_only=True)
    time.sleep(10)
    sheet = file.active

    result = []

    try:
        for item in data:
            row_data = {
                "id": item['id'],
                "codClient": item['codClient'],
                "nameClient": item['nameClient'],
                "codNewProd": item['codNewProd'],
                "newProduct": item['newProduct'],
                "grid": item['grid'],    
                "seller": item['seller'],
                "history":[]
            }
            # ALTERAR COD PARA DESC_PRODUTO - X
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (str(row[0]) == item['codClient'] or str(row[1]) == item['nameClient']):
                    history = {
                        "cod": str(row[23]),
                        "grid": str(row[69]),    
                        "seller": str(row[71]),
                        "data": str(row[14])
                    }
                    row_data['history'].append(history)

            result.append(row_data)
                    
    except Exception as e:
        logger(f"!! O script encontrou um erro (getHistory) : {e}")

    file.close()

    logger(f"Historicos de clientes carregado com sucesso: {str(result)}")

    if not result:
        data

    return result

def verifyLastContact(data):
    try:
        #-- ABRE PLANILHA
        file = openpyxl.load_workbook(SHEET_NOVOS_PRODUTOS)
        sheet = file['Enviado']  

        result = []
        #-- PERCORRE ARRAY
        for row_data in data:
            matching_row_found = False

            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[2] == row_data['codClient']:
                    matching_row_found = True
                    lastContact = parser.parse(row[0], dayfirst=True)
                    diferenca_em_dias = (datetime.now() - lastContact).days

                    logger(f"diferenca_em_dias: {diferenca_em_dias}")
                    logger(f"row_data[prazo]: {row_data['prazo']}")

                    if diferenca_em_dias >= row_data["prazo"]:
                        result.append(row_data)

            # Se nenhum correspondência foi encontrada, adiciona row_data ao resultado
            if not matching_row_found:
                result.append(row_data)

    except Exception as e:
        logger(f"!! O script encontrou um erro (verifyLastContact): {e}")
        return False
    
    finally:
        file.close()
    
    return result

def getSeller(seller):
    file = openpyxl.load_workbook(SHEET_VENDEDORES, data_only=True)
    time.sleep(3)
    sheet = file.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if (str(row[0]) == seller):
            return str(row[1])
    
    return False

def alterLastContact(item) -> bool:
    try:
        file = openpyxl.load_workbook(SHEET_NOVOS_PRODUTOS)
        sheet = file['Enviado']  

        data_atual = datetime.now()
        datastring = data_atual.strftime("%d/%m/%Y")

        # Verifica se já existe uma linha com o código do cliente
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == item['codClient']:
                sheet.cell(row=row[0], column=1, value=datastring)
                sheet.cell(row=row[0], column=2, value=item['newProduct'])
                sheet.cell(row=row[0], column=3, value=item['codClient'])
                sheet.cell(row=row[0], column=4, value=item['nameClient'])
                break
        else:
            # Se não encontrou uma linha com o código do cliente, adiciona uma nova linha no início da planilha
            new_row_index = 2
            sheet.insert_rows(new_row_index)

            # Escreve os dados na nova linha
            sheet.cell(row=new_row_index, column=1, value=datastring)
            sheet.cell(row=new_row_index, column=2, value=item['newProduct'])
            sheet.cell(row=new_row_index, column=3, value=item['codClient'])
            sheet.cell(row=new_row_index, column=4, value=item['nameClient'])

    except Exception as e:
        logger(f"!! O script encontrou um erro (alterLastContact): {e}")
        return False
    
    finally:
        file.save(SHEET_NOVOS_PRODUTOS) 
        file.close()
    
    return True
